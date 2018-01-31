import os
import shutil
import openpyxl
import re


# source directory - generated screenshots are stored here
dir_src = os.path.expanduser("~/collect/fastlane/metadata/android/en/images/phoneScreenshots/")

# destination directory
utilPath=(os.path.dirname(os.path.realpath(__file__)))
xlsPath=utilPath+"/all-widgets.xlsx"
xmlPath= utilPath+"/all-widgets.xml"

# pyxform conversion
os.system("xls2xform " + xlsPath + " " + xmlPath)

pathList= utilPath.split("/")
pathList.remove(pathList[-1])
fullPath="/".join(pathList)
fullImgPath= fullPath +"/img/form-widgets" #path to form-widgets image directory
dir_dst = fullImgPath

path=dir_dst

# delete the existing images
for the_file in os.listdir(path):
    file_path = os.path.join(path, the_file)
    try:
        if os.path.isfile(file_path):
            os.unlink(file_path)
    except Exception as e:
        print(e)


for file in os.listdir(dir_src):
    print(file)  # testing
    src_file = os.path.join(dir_src, file)
    dst_file = os.path.join(dir_dst, file)
    shutil.move(src_file, dst_file)

files = os.listdir(path)

# To remove unwanted numbers appended to screenshots' names
for file in files:
    if "_" in file:
        f=file.split("_")
        os.rename(os.path.join(path, file), os.path.join(path, f[0]+'.png'))

# path to form-widgets guide
fullFilepath = fullPath + "/form-widgets.rst" 

''' Extracts introductory sentences'''
intro = []
lookup=" section shows examples of all the form widgets types"

with open(fullFilepath) as file:
	for num, line in enumerate(file, 1):
		if lookup in line:
			found=num
			with open(fullFilepath) as file:
				intro.extend(file.readline() for i in range(found))


widgets = open(fullFilepath, "w")
widgets.seek(0)
#prints intro
widgets.writelines(intro)
widgets.write("\n")


xmlFile = open("all-widgets.xml", "r+")
text = xmlFile.read()

wb= openpyxl.load_workbook('all-widgets.xlsx')
sheet=wb['survey']

# prints XLSForm Rows
def rows(row_index):
	widgets.write(sheet.cell(row = 1, column=10).value + "\n\n")
	widgets.write(".. csv-table:: survey" +"\n")
	widgets.write(" :header: type, name, label"+"\n")
	values=[]
	if ((sheet.cell(row= row_index, column=1).value) !="begin_group" and (sheet.cell(row= row_index, column=1).value) !="end_group" ):
		if (sheet.cell(row= row_index, column=1).value is not None):
			# appends type of the widget
			values.append(sheet.cell(row=row_index, column=1).value)
		if (sheet.cell(row= row_index, column=3).value is not None):
			# append name of the widget
			values.append(sheet.cell(row=row_index, column= 3).value)
		if (sheet.cell(row= row_index, column=4).value is not None):
			# appends label of the widget
			values.append(sheet.cell(row=row_index, column= 4).value)
		widgets.write(" "+ ", ".join(values))
		widgets.write("\n\n")
		if len(values)==3:
			# values[1]= name of widget
			xml(values[1])

# prints XForm XML
def xml(name):
	widgets.write(sheet.cell(row = 1, column=11).value + "\n\n")
	widgets.write(".. code-block:: xml"+ "\n\n")
	with open('all-widgets.xml', 'rt') as file:
		for line in file:
			if "bind" in line:
				if name in line:
					xmlSplit=(line.split("/"))
					nameSplit=(xmlSplit[3].split(" "))
					length=len(nameSplit[0])
					widgetName=nameSplit[0]
					if (widgetName[0:length-1]==name):
						widgets.write("  "+line.strip()+"\n")

			if "input" in line:
				if name in line:
					snippet=line.strip()
					snippetSplit=snippet.split("/")
					snippetName=snippetSplit[-1]
					snipLength=len(snippetName)
					if (snippetName[0:snipLength-2]== name):
						widgets.write("  "+snippet+"\n")
						# extracts elements between input tags
						expression =[x.strip() for x in re.findall(snippet+'(.*?)</input>', text, re.MULTILINE | re.DOTALL)]
						widgets.write("  "+"".join(expression)+ "\n")
						widgets.write("  "+"</input>"+"\n\n")


for i in range(4,86):
	# prints headlines and section labels
	if (sheet.cell(row=i, column=2).value is not None):
		widgets.writelines(sheet.cell(row=i, column=2).value + "\n\n")
	# prints captions
	if (sheet.cell(row=i, column=7).value is not None):
		widgets.writelines(sheet.cell(row=i, column=7).value + "\n")
		widgets.writelines("\n")
	# prints image directive
	if (sheet.cell(row=i, column=9).value is not None):	
		widgets.writelines(".. image:: /img/widgets/"+ sheet.cell(row=i, column=9).value +".*\n")
	# prints alt texts
	if (sheet.cell(row=i, column=8).value is not None):
		widgets.writelines(" alt:"+ sheet.cell(row=i, column=8).value+"\n")
	widgets.write("\n")
	rows(i)




  

